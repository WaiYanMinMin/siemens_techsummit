#!/usr/bin/env python3
"""Generate Kaspersky APAC Partner Conference 2026 cue sheets from agenda HTML."""

from __future__ import annotations

import json
import re
from copy import copy
from datetime import datetime, time, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HTML_PATH = Path("/Users/waiyanminmin/Downloads/kaspersky-conference-2026 (3).html")
TEMPLATE_PATH = Path(
    "/Users/waiyanminmin/Downloads/Kaspersky 2026_ Emcee script_El_AGENDA_UPDATED.xlsx"
)
OUTPUT_PATH = Path(
    "/Users/waiyanminmin/Downloads/Kaspersky_APAC_Partner_Conference_2026_Cue_Sheets.xlsx"
)

HEADERS = [
    "S/N",
    "Start Time",
    "End Time",
    "Dur.",
    "Agenda",
    "in-person Speakers and Designation",
    "Programme Remarks",
    "Emcee Script",
    "Preset",
    "LED Wall (16:9) (6mW x 3.5mH)",
    "Side Walls (LEFT AND RIGHT) (2mW x 3.5mH)",
    "Preset",
    "Live Stream Main + PIP (16:9)",
    "Online Speakers and Designation",
    "Sound",
    "Stage / Crew",
    "Light",
    "Mic",
]

TIME_FMT = "h:mm:ss"
DUR_FMT = "h:mm:ss"


def load_agendas() -> dict:
    text = HTML_PATH.read_text(encoding="utf-8")
    match = re.search(r"const AGENDAS = (\{[\s\S]*?\});", text)
    if not match:
        raise RuntimeError("AGENDAS block not found in HTML")
    return json.loads(match.group(1))


def parse_clock(value: str) -> time | None:
    value = value.strip().upper()
    for fmt in ("%H:%M", "%I:%M %p", "%I:%M%p"):
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    return None


def parse_time_range(raw: str) -> tuple[time | None, time | None]:
    if not raw or raw.lower().startswith("all day") or raw.lower().startswith("by "):
        return None, None
    normalized = (
        raw.replace("–", "-")
        .replace("—", "-")
        .replace("−", "-")
        .strip()
    )
    parts = [p.strip() for p in normalized.split("-") if p.strip()]
    if len(parts) != 2:
        return None, None
    start = parse_clock(parts[0])
    end = parse_clock(parts[1])
    return start, end


def to_excel_time(t: time | None):
    return t


def duration(start: time | None, end: time | None):
    if not start or not end:
        return None
    s = datetime.combine(datetime.today(), start)
    e = datetime.combine(datetime.today(), end)
    if e <= s:
        e += timedelta(minutes=1)
    return e - s


def bump_minute(t: time | None, minutes: int = 1) -> time | None:
    if t is None:
        return None
    dt = datetime.combine(datetime.today(), t) + timedelta(minutes=minutes)
    return dt.time()


def speaker_line(item: dict) -> str:
    parts = []
    spk = (item.get("speaker") or "").strip()
    loc = (item.get("location") or "").strip()
    if spk:
        parts.append(spk)
    if loc and loc not in parts:
        if loc.lower().startswith("salon"):
            pass
        elif not spk or any(
            token in loc for token in (",", "·", "CEO", "VP", "Director", "Managing")
        ):
            parts.append(loc)
    return " · ".join(parts) if parts else ""


def remarks(item: dict) -> str:
    notes = (item.get("notes") or "").strip()
    loc = (item.get("location") or "").strip()
    if loc.lower().startswith("salon"):
        venue = f"Venue: {loc}"
        return f"{venue} · {notes}" if notes else venue
    return notes


def intro_script(session: str, speaker: str) -> str:
    if not speaker:
        return ""
    return f"Please welcome {speaker} for {session}."


def outtro_script(session: str, speaker: str) -> str:
    if not speaker:
        return f"Thank you. That concludes {session}."
    name = speaker.split(",")[0].split("·")[0].strip()
    return f"Thank you, {name}, for {session}."


def production_defaults(kind: str) -> dict:
    if kind == "registration":
        return {
            "preset": "1.0",
            "led": "Loop video / holding slide",
            "walls": "Holding animated video on loop",
            "live_preset": "1.0",
            "live": "16:9 Holding",
            "sound": "BGM ON",
            "stage": "Registration team ready",
        }
    if kind == "opening":
        return {
            "preset": "3.0",
            "led": "Housekeeping slides",
            "walls": "Holding Video",
            "live_preset": "2.0",
            "live": "N/A — in-person partner conference",
            "sound": "Walk-in OUT, Mic ON",
            "stage": "Cue emcee",
        }
    if kind == "session":
        return {
            "preset": "3.0",
            "led": "PPT",
            "walls": "Live-feed",
            "live_preset": "3.0",
            "live": "N/A",
            "sound": "Walk-in sting",
            "stage": "Escort speaker",
        }
    if kind == "panel":
        return {
            "preset": "3.0",
            "led": "Panel lower-thirds / PPT",
            "walls": "Live-feed",
            "live_preset": "3.0",
            "live": "N/A",
            "sound": "Panel mics ON",
            "stage": "Moderator + panelists seated",
        }
    if kind == "break":
        return {
            "preset": None,
            "led": "Holding slide — Kaspersky branding",
            "walls": "Holding slide",
            "live_preset": "1.0",
            "live": "16:9 Holding",
            "sound": "BGM ON",
            "stage": "Reset stage / refresh breakout rooms",
        }
    if kind == "breakout":
        return {
            "preset": "3.0",
            "led": "PPT",
            "walls": "Live-feed (if AV in salon)",
            "live_preset": "3.0",
            "live": "N/A",
            "sound": "Mic ON",
            "stage": "Salon AV team standby",
        }
    if kind == "gala":
        return {
            "preset": None,
            "led": "Gala holding / awards visuals",
            "walls": "Ambient / awards loop",
            "live_preset": "1.0",
            "live": "N/A",
            "sound": "BGM ON",
            "stage": "Banquet ops / awards stage manager",
        }
    return {}


def apply_prod(row: dict, kind: str):
    prod = production_defaults(kind)
    row.update(
        {
            "preset": prod.get("preset"),
            "led": prod.get("led"),
            "walls": prod.get("walls"),
            "live_preset": prod.get("live_preset"),
            "live": prod.get("live"),
            "online_speaker": row.get("online_speaker"),
            "sound": prod.get("sound"),
            "stage": prod.get("stage"),
        }
    )


def make_row(
    *,
    sn: float | None,
    agenda: str,
    start: time | None,
    end: time | None,
    speaker: str = "",
    programme: str = "",
    emcee: str = "",
    kind: str = "session",
) -> dict:
    row = {
        "sn": sn,
        "start": start,
        "end": end,
        "dur": duration(start, end),
        "agenda": agenda,
        "speaker": speaker,
        "programme": programme,
        "emcee": emcee,
        "online_speaker": speaker if speaker else None,
    }
    apply_prod(row, kind)
    return row


def expand_session(item: dict, sn_counter: list, kind: str = "session") -> list[dict]:
    start, end = parse_time_range(item["time"])
    session = item["session"]
    spk = speaker_line(item)
    prog = remarks(item)
    rows: list[dict] = []

    skip_intro = session.lower() in {
        "registration & welcome",
        "lunch",
        "tea break",
        "rest & refresh",
        "rest",
        "b2c break-out",
        "eu break-out",
    }

    if not skip_intro and spk and session not in {"Break", "Opening Remarks"}:
        intro_end = bump_minute(start)
        rows.append(
            make_row(
                sn=4.0,
                agenda="Speaker Intro",
                start=start,
                end=intro_end,
                speaker=spk,
                programme=prog or "Speaker introduction",
                emcee=intro_script(session, spk),
                kind=kind,
            )
        )
        if intro_end and start and intro_end > start:
            start = intro_end

    sn_counter[0] += 1
    rows.append(
        make_row(
            sn=float(sn_counter[0]),
            agenda=session,
            start=start,
            end=end,
            speaker=spk,
            programme=prog or session,
            emcee="",
            kind=kind,
        )
    )

    if not skip_intro and spk and session.lower() not in {"lunch", "tea break", "break"}:
        out_start = end
        out_end = bump_minute(out_start)
        rows.append(
            make_row(
                sn=6.0,
                agenda="Outtro",
                start=out_start,
                end=out_end,
                speaker=spk,
                programme=prog or "Transition",
                emcee=outtro_script(session, spk),
                kind=kind,
            )
        )
    return rows


def get_main_conference_day(agendas: dict) -> dict:
    """Plenary is identical across tracks; prefer APC for breakout metadata."""
    for key in ("APC", "EU", "PAC"):
        if key not in agendas:
            continue
        for day in agendas[key]["days"]:
            if "MAIN CONFERENCE" in day["title"]:
                return day
    for track_data in agendas.values():
        for day in track_data["days"]:
            if "MAIN CONFERENCE" in day["title"]:
                return day
    raise RuntimeError("Main conference day not found")


def plenary_items(day: dict) -> list[dict]:
    for sub in day["subsections"]:
        if "Plenary" in sub["title"] or "All Attendees" in sub["title"]:
            return sub["items"]
    raise RuntimeError("Plenary subsection not found")


def eu_breakout_items(agendas: dict) -> list[dict]:
    for key in ("EU", "APC", "PAC"):
        if key not in agendas:
            continue
        for day in agendas[key]["days"]:
            if "MAIN CONFERENCE" not in day["title"]:
                continue
            for sub in day["subsections"]:
                if "EU Breakout" in sub["title"]:
                    return [
                        i
                        for i in sub["items"]
                        if i["session"] not in {"Rest & Refresh"}
                    ]
    return []


def apac_b2b_items(day: dict) -> list[dict]:
    skip = {"B2C Break-out", "EU Break-out", "Rest"}
    items = []
    for sub in day["subsections"]:
        if "APAC B2B" in sub["title"]:
            items.extend(sub["items"])
    return [i for i in items if i["session"] not in skip]


def gala_items(day: dict) -> list[dict]:
    gala = []
    for sub in day["subsections"]:
        if "Gala" in sub["title"]:
            gala.extend(sub["items"])
    if not gala:
        return [
            {
                "time": "18:30 – 19:00",
                "session": "Guest Arrival & Seating",
                "speaker": "",
                "location": "JW Marriott Seoul — Gala venue (TBC)",
                "notes": "Guests assemble; BGM; emcee standby",
            },
            {
                "time": "19:00 – 19:15",
                "session": "Welcome & Opening Remarks",
                "speaker": "Kaspersky leadership (TBC)",
                "location": "",
                "notes": "Official welcome to Gala Dinner & Award Ceremony",
            },
            {
                "time": "19:15 – 20:15",
                "session": "Dinner Service",
                "speaker": "",
                "location": "",
                "notes": "Service flow per banquet ops",
            },
            {
                "time": "20:15 – 21:30",
                "session": "Award Ceremony",
                "speaker": "Presenters TBC",
                "location": "",
                "notes": "Partner awards — confirm recipient list with events team",
            },
            {
                "time": "21:30 – 22:00",
                "session": "Closing & Farewell",
                "speaker": "",
                "location": "",
                "notes": "Thank guests; announce 11 June DMZ reminder (passport + Kaspersky tee)",
            },
        ]
    expanded = []
    for item in gala:
        if "19:00" in item["time"] and "22:00" in item["time"]:
            expanded.extend(
                [
                    {
                        "time": "18:30 – 19:00",
                        "session": "Guest Arrival & Seating",
                        "speaker": "",
                        "location": item.get("location", ""),
                        "notes": "Pre-function; guests to assigned tables",
                    },
                    {
                        "time": "19:00 – 19:15",
                        "session": "Welcome Remarks",
                        "speaker": "Kaspersky leadership (TBC)",
                        "location": "",
                        "notes": item["session"],
                    },
                    {
                        "time": "19:15 – 20:15",
                        "session": "Dinner Service",
                        "speaker": "",
                        "location": "",
                        "notes": "Per banquet run-of-show",
                    },
                    {
                        "time": "20:15 – 21:30",
                        "session": "Award Ceremony",
                        "speaker": "Presenters TBC",
                        "location": "",
                        "notes": "Partner awards presentation",
                    },
                    {
                        "time": "21:30 – 22:00",
                        "session": "Closing",
                        "speaker": "",
                        "location": "",
                        "notes": item.get("notes", ""),
                    },
                ]
            )
        else:
            expanded.append(item)
    return expanded


def b2c_items() -> list[dict]:
    return [
        {
            "time": "13:50 – 14:00",
            "session": "B2C Breakout — Welcome & Housekeeping",
            "speaker": "Track lead TBC",
            "location": "Salon 3, Level 3",
            "notes": "Detailed B2C session list not published in partner agenda microsite — confirm with programme team",
        },
        {
            "time": "14:00 – 16:25",
            "session": "B2C Breakout Programme Block",
            "speaker": "Speakers TBC",
            "location": "Salon 3, Level 3",
            "notes": "Parallel to APAC B2B (Ballroom track) and EU Breakout (Salon 1); ends 16:25 per master agenda",
        },
    ]


def build_sheet_rows(items: list[dict], default_kind: str) -> list[dict]:
    rows: list[dict] = []
    sn = [0]

    prep = make_row(
        sn=None,
        agenda="JW Marriott Seoul · Level 5 Ballroom (Plenary) / Level 3 Salons (Breakouts) · 10 June 2026",
        start=time(8, 30),
        end=None,
        programme="Cue sheet generated from partner agenda microsite",
        kind=default_kind,
    )
    rows.append(prep)

    for item in items:
        session = item["session"].lower()
        if "registration" in session:
            kind = "registration"
        elif session in {"lunch", "tea break"}:
            kind = "break"
        elif "opening" in session or "welcome" in session and "registration" not in session:
            kind = "opening"
        elif "panel" in session or "sharing" in session:
            kind = "panel"
        elif default_kind == "gala":
            kind = "gala"
        elif default_kind == "breakout":
            kind = "breakout"
        else:
            kind = "session"

        rows.extend(expand_session(item, sn, kind=kind))
    return rows


def write_rows(ws, rows: list[dict], template_ws):
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if template_ws.cell(1, col_idx).has_style:
            cell.font = copy(template_ws.cell(1, col_idx).font)
            cell.fill = copy(template_ws.cell(1, col_idx).fill)
            cell.border = copy(template_ws.cell(1, col_idx).border)
            cell.alignment = copy(template_ws.cell(1, col_idx).alignment)

    r = 2
    for row in rows:
        ws.cell(r, 1, row.get("sn"))
        for col, key in [(2, "start"), (3, "end"), (4, "dur")]:
            val = row.get(key)
            if val is None:
                continue
            if isinstance(val, timedelta):
                ws.cell(r, col, val)
                ws.cell(r, col).number_format = DUR_FMT
            elif isinstance(val, time):
                ws.cell(r, col, to_excel_time(val))
                ws.cell(r, col).number_format = TIME_FMT
            else:
                ws.cell(r, col, val)
        ws.cell(r, 5, row.get("agenda"))
        ws.cell(r, 6, row.get("speaker"))
        ws.cell(r, 7, row.get("programme"))
        ws.cell(r, 8, row.get("emcee"))
        ws.cell(r, 9, row.get("preset"))
        ws.cell(r, 10, row.get("led"))
        ws.cell(r, 11, row.get("walls"))
        ws.cell(r, 12, row.get("live_preset"))
        ws.cell(r, 13, row.get("live"))
        ws.cell(r, 14, row.get("online_speaker"))
        ws.cell(r, 15, row.get("sound"))
        ws.cell(r, 16, row.get("stage"))
        ws.cell(r, 17, None)
        ws.cell(r, 18, None)
        r += 1


def copy_sheet_setup(ws, template_ws):
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = template_ws.column_dimensions[letter].width
    ws.freeze_panes = "A2"


def main():
    agendas = load_agendas()
    day = get_main_conference_day(agendas)

    wb = load_workbook(TEMPLATE_PATH)
    template_ws = wb["Emcee_Agenda_Updated"]

    for name in list(wb.sheetnames):
        del wb[name]

    sheets = [
        (
            "1_Plenary_All_Attendees",
            build_sheet_rows(plenary_items(day), "session"),
        ),
        (
            "2_Breakout_EU_Salon1",
            build_sheet_rows(eu_breakout_items(agendas), "breakout"),
        ),
        (
            "3_Breakout_APAC_B2B",
            build_sheet_rows(apac_b2b_items(day), "breakout"),
        ),
        (
            "4_Breakout_B2C_Salon3",
            build_sheet_rows(b2c_items(), "breakout"),
        ),
        (
            "5_Gala_Dinner_Awards",
            build_sheet_rows(gala_items(day), "gala"),
        ),
    ]

    for title, rows in sheets:
        ws = wb.create_sheet(title)
        copy_sheet_setup(ws, template_ws)
        write_rows(ws, rows, template_ws)

    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    for title, rows in sheets:
        print(f"  {title}: {len(rows)} cue rows")


if __name__ == "__main__":
    main()
